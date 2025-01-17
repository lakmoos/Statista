import os
import pandas as pd
import logging
import warnings
import openpyxl
from openpyxl import Workbook
import re


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("scraper.log", mode="w", encoding="utf-8"),
    ],
)
log = logging.getLogger()


def format_relative_path(base_dir, full_path):
    """Format and return the relative path from the base directory."""
    try:
        return os.path.relpath(full_path, base_dir)
    except ValueError:
        return full_path


def process_files(selected_files, base_dir, process_function, skip_adv=False):
    """
    Generic function to process selected files using a given transformation function.
    Optionally skips files containing "adv" in their name if skip_adv is True.
    """
    transformed_files = []
    errors = []

    for relative_path in selected_files:
        try:
            input_path = os.path.join(base_dir, relative_path)

            if not os.path.exists(input_path):
                logging.warning(
                    f"File not found: {relative_path}. Skipping transformation."
                )
                errors.append(relative_path)
                continue

            if skip_adv and "adv" in os.path.basename(input_path).lower():
                logging.info(
                    f"Skipping file {relative_path} as it contains 'adv' in the name."
                )
                transformed_files.append(format_relative_path(base_dir, input_path))
                continue

            # Determine the transformed output directory
            parent_dir = os.path.dirname(input_path)
            transformed_dir = (
                parent_dir
                if parent_dir.endswith("transformed")
                else os.path.join(parent_dir, "transformed")
            )
            os.makedirs(transformed_dir, exist_ok=True)

            output_path = os.path.join(transformed_dir, os.path.basename(input_path))
            formatted_path = format_relative_path(base_dir, output_path)

            # Process the file using the provided function
            with pd.ExcelFile(input_path) as xls:
                sheet_data = process_function(xls)

            # If we are dealing with a single workbook - when merging
            if isinstance(sheet_data, Workbook):
                logging.info("Saving workbook.")
                sheet_data.save(output_path)
            else:
                # Save the processed data
                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    if isinstance(sheet_data, dict):
                        for sheet_name, data in sheet_data.items():
                            data.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        logging.error(
                            f"An error occurred while saving file {relative_path}: Bad file type."
                        )
                        errors.append(relative_path)

            transformed_files.append(formatted_path)

        except Exception as e:
            logging.error(
                f"An error occurred while processing file {relative_path}: {e}"
            )
            errors.append(relative_path)

    # Log transformation summary
    logging.info("Selected files to transform: %d", len(selected_files))
    if transformed_files:
        logging.info("Successfully saved:")
        for path in transformed_files:
            logging.info(f"    {path}")
    if errors:
        logging.error("Errors occurred during transformation for the following files:")
        for path in errors:
            logging.error(f"    {path}")

    return transformed_files


def tr1_remove_sheets(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheets = xls.sheet_names
        return {
            sheet: pd.read_excel(xls, sheet_name=sheet, header=None)
            for sheet in sheets
            if sheet not in ["Overview", "Content", "Lists"]
        }

    return process_files(selected_files, base_dir, process_function)


def tr2_remove_header_and_empty_column(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Remove metadata rows (rows without numeric data)
            valid_data_start_index = df[
                df.apply(
                    lambda col: col.map(lambda x: isinstance(x, (int, float))), axis=0
                ).any(axis=1)
            ].index.min()

            df_cleaned = df.iloc[valid_data_start_index:].reset_index(drop=True)

            # Use the first valid row as headers
            df_cleaned.columns = df_cleaned.iloc[0]
            df_cleaned = df_cleaned[1:].reset_index(drop=True)

            # Drop entirely empty columns
            df_cleaned = df_cleaned.dropna(axis=1, how="all")

            # Replace NaN or placeholder headers
            df_cleaned.columns = [
                f"Column_{i}" if pd.isna(col) else col
                for i, col in enumerate(df_cleaned.columns)
            ]

            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function, skip_adv=True)


def tr3_remove_metadata(selected_files, base_dir="statista_data"):
    keywords = [
        "Survey Name:",
        "Base n =",
        "Question Type:",
        "Sample Size n =",
        "Population:",
        "¹Low base:",
        "Survey period",
        "Survey name:",
        "Sample size n = ",
    ]

    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df_cleaned = df[
                ~df.apply(
                    lambda row: row.astype(str)
                    .str.contains("|".join(keywords), na=False)
                    .any(),
                    axis=1,
                )
            ]
            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr4_reduce_empty_lines(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Replace consecutive empty rows with a single empty row
            df_cleaned = df.loc[
                (df.shift(1).isnull().all(axis=1) & df.isnull().all(axis=1)) == False
            ]

            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr5_removing_total_percentages(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Remove columns where any cell contains "Grand Total" or "in %"
            columns_to_keep = df.apply(
                lambda col: ~col.astype(str)
                .str.contains("Grand Total|in %", na=False)
                .any()
            )
            df_cleaned = df.loc[:, columns_to_keep]

            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr6_append_questions(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}

        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df.reset_index(drop=True, inplace=True)

            question = None
            rows_to_drop = []

            # Process the DataFrame for the current sheet
            for idx, row in df.iterrows():
                first_col = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""

                # Identify and extract question rows
                match = re.match(r"^(.*?\?).*", first_col)
                if match:
                    question = match.group(1).strip()  # Extract question until `?`
                    rows_to_drop.append(idx)  # Collect the index of the question row
                    continue

                # Append question to the first column of option rows
                if question and first_col:
                    df.iloc[idx, 0] = f"{question} {first_col}"

            # Drop question rows
            df.drop(index=rows_to_drop, inplace=True)
            df.reset_index(drop=True, inplace=True)

            # Add the processed DataFrame to the output dictionary
            sheet_data[sheet] = df

        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr7_merging_sheets(selected_files, base_dir="statista_data"):
    def process_function(xls):
        wb = openpyxl.load_workbook(xls)
        new_wb = Workbook()
        merged_sheet = new_wb.active
        merged_sheet.title = "Merged Data"
        current_row = 1

        for sheet_name in wb.sheetnames:

            sheet = wb[sheet_name]

            # Copy rows from the current sheet to the merged sheet
            for row in sheet.iter_rows(values_only=True):
                merged_sheet.append(row)

            # Add a blank row between sheets (optional for better readability)
            current_row += sheet.max_row
            merged_sheet.append([])

        # FIXME: the first row seems useless, so it's dropped - just in case
        merged_sheet.delete_rows(1)
        return new_wb

    return process_files(selected_files, base_dir, process_function)


def pipeline_transform(selected_files, base_dir="statista_data"):
    logging.info("Starting the transformation pipeline...")

    # Step 1: Remove Overview
    logging.info("Step 1: Removing Overview sheets...")
    transformed_step1 = tr1_remove_sheets(selected_files, base_dir)
    if not transformed_step1:
        logging.warning("No files to process after Step 1. Exiting pipeline.")
        return

    # Step 2: Remove header and empty columns
    logging.info("Step 2: Removing header rows and empty columns...")
    transformed_step2 = tr2_remove_header_and_empty_column(transformed_step1, base_dir)
    if not transformed_step2:
        logging.warning("No files to process after Step 2. Exiting pipeline.")
        return

    # Step 3: Remove metadata rows
    logging.info("Step 3: Removing metadata rows...")
    transformed_step3 = tr3_remove_metadata(transformed_step2, base_dir)

    # Step 4: Reduce empty lines
    logging.info("Step 4: Reducing empty lines...")
    transformed_step4 = tr4_reduce_empty_lines(transformed_step3, base_dir)

    # # Step 5: Remove total percentages columns
    logging.info("Step 5: Removing columns with 'Grand Total' and 'in %'...")
    transformed_step5 = tr5_removing_total_percentages(transformed_step4, base_dir)

    # Step 6: Append questions to options
    logging.info("Step 6: Appending questions to options...")
    transformed_step6 = tr6_append_questions(transformed_step5, base_dir)

    # # Step 7: Merge sheets
    logging.info("Step 6: Merging sheets...")
    transformed_step7 = tr7_merging_sheets(transformed_step6, base_dir)

    logging.info("Transformation pipeline completed.")
    return transformed_step7
